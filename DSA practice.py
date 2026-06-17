# import openpyxl as xl
# file = xl.load_workbook('STOCK_18%(1).xlsx')
# sheet = file['Stock Summary']

# for row in range(4,sheet.max_row+1):
#     cell_1=sheet.cell(row,6)
#     cell_2=sheet.cell(row,7)
    
#     value_1 = cell_1.value if cell_1.value is not None else 0
#     value_2 = cell_2.value if cell_2.value is not None else 0
#     total_price = value_1 + value_2
    
#     sheet.cell(row,8).value=total_price
    
    
    
# file.save('STOCK_18%(1).xlsx')
    
import openpyxl as xl

# Load the workbook and select the sheet
file = xl.load_workbook('STOCK_18%(1).xlsx')
sheet = file['Stock Summary']

# Loop through rows starting from row 4
for row in range(3, sheet.max_row + 1):
    sheet.cell(row,4).value = f"=B{row}*C{row}"

file.save('STOCK_18%(1).xlsx')